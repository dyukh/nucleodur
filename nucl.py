import pymupdf  # PyMuPDF
from openpyxl import load_workbook


def extract_stream_from_pdf(pdf_path):
    doc = pymupdf.open(stream=pdf_path, filetype="pdf")
    for page in doc:
        for xref in page.get_contents():
            stream = doc.xref_stream(xref).decode("latin1")
            # if "stream" in stream and "endstream" in stream:
            # Извлекаем данные между stream и endstream
            data = stream.split("stream")[0].split("endstream")[0].strip()
            start = data.find("W n")
            stop = data[start:].find("S")
            return data[start: start + stop - 1]
    return None


def parse_path_data(data):
    commands = []
    tokens = data.split()
    i = 0
    while i < len(tokens):
        if tokens[i].isalpha():  # Команда (m, l, c...)
            cmd = tokens[i]
            i += 1
            points = []
            while i + 1 < len(tokens) and not tokens[i].isalpha():
                x = float(tokens[i])
                y = float(tokens[i + 1])
                points.append((x, y))
                i += 2
            commands.append((cmd, points))
        else:
            i += 1
    return commands


def to_list(commands):
    plist = []
    for cmd, points in commands:
        if points and len(points[0]) > 1:
            plist.append([points[0][1], points[0][0]])
    return plist


def to_excel(plist, fname="out.xlsx", template="template.xlsx"):
    # sheet_name = "Sheet1"
    wb = load_workbook(template)
    ws = wb.active

    # Обновление данных в шаблоне
    for i, row in enumerate(plist):
        # ws.append(row)
        ws.cell(row=i+2, column=1).value = row[0]
        ws.cell(row=i+2, column=2).value = row[1]
    # Сохранение обновленного файла
    wb.save(fname)


if __name__ == "__main__":
    pdf_path = "test.pdf"
    stream_data = extract_stream_from_pdf(pdf_path)
    path_commands = parse_path_data(stream_data)
    plist = to_list(path_commands)
    to_excel(plist)
